import base64
import hashlib
import html
import io
import json
import logging
import os
import re
import sqlite3
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from anthropic import Anthropic
from flask import Flask, jsonify, request
from openpyxl.styles import Alignment, Font, PatternFill
from pypdf import PdfReader


logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
logger = logging.getLogger("metra_accounting")

BOT_TOKEN = os.environ["BOT_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
PUBLIC_URL = os.getenv("PUBLIC_URL", "").rstrip("/")
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "")
SETUP_SECRET = os.getenv("SETUP_SECRET", "")
ALLOWED_USERS = {
    int(value.strip())
    for value in os.getenv("ALLOWED_TELEGRAM_USER_IDS", "").split(",")
    if value.strip().isdigit()
}

if not ALLOWED_USERS:
    raise RuntimeError("ALLOWED_TELEGRAM_USER_IDS is required")

DATA_DIR = Path(os.getenv("DATA_DIR", "/data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
RECEIPT_DIR = DATA_DIR / "receipts"
RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
STATEMENT_DIR = DATA_DIR / "card_statements"
STATEMENT_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "accounting.db"

CARD_BASELINE_MONTH = os.getenv("CARD_BASELINE_MONTH", "2026-04")
CARD_MONTHLY_BUDGET = float(os.getenv("CARD_MONTHLY_BUDGET", "5000"))

app = Flask(__name__)
executor = ThreadPoolExecutor(max_workers=4)

COMPANY_CATEGORIES = [
    "Transport et véhicule",
    "Repas et représentation",
    "Bureau et loyer",
    "Technologie et logiciels",
    "Télécom et internet",
    "Matériel et équipement",
    "Services professionnels",
    "Marketing et publicité",
    "Voyage et déplacement",
    "Formation",
    "Assurances",
    "Taxes et licences",
    "Fournitures de bureau",
    "Sous-traitance",
    "Autre dépense",
]

PERSONAL_CATEGORIES = [
    "Épicerie et alimentation",
    "Restaurants et sorties",
    "Transport et essence",
    "Logement et services",
    "Vêtements",
    "Santé et médical",
    "Loisirs",
    "Abonnements et technologie",
    "Voyage",
    "Cadeaux et dons",
    "Autre dépense",
]


def db():
    connection = sqlite3.connect(DB_PATH, timeout=20)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    return connection


def init_db():
    with db() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                user_id INTEGER PRIMARY KEY,
                step TEXT,
                payload TEXT NOT NULL DEFAULT '{}',
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                merchant TEXT NOT NULL,
                expense_date TEXT NOT NULL,
                subtotal REAL NOT NULL DEFAULT 0,
                gst REAL NOT NULL DEFAULT 0,
                qst REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL DEFAULT 0,
                currency TEXT NOT NULL DEFAULT 'CAD',
                description TEXT,
                expense_type TEXT NOT NULL,
                category TEXT NOT NULL,
                project_code TEXT,
                receipt_hash TEXT NOT NULL,
                receipt_path TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, receipt_hash)
            );

            CREATE TABLE IF NOT EXISTS card_statements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                statement_date TEXT NOT NULL,
                period_start TEXT,
                period_end TEXT,
                previous_balance REAL NOT NULL DEFAULT 0,
                payments_credits REAL NOT NULL DEFAULT 0,
                purchases_debits REAL NOT NULL DEFAULT 0,
                cash_advances REAL NOT NULL DEFAULT 0,
                interest REAL NOT NULL DEFAULT 0,
                fees REAL NOT NULL DEFAULT 0,
                ending_balance REAL NOT NULL DEFAULT 0,
                minimum_payment REAL NOT NULL DEFAULT 0,
                payment_due_date TEXT,
                credit_limit REAL NOT NULL DEFAULT 0,
                card_last4 TEXT,
                source_hash TEXT NOT NULL,
                source_path TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, source_hash),
                UNIQUE(user_id, statement_date, card_last4)
            );
            """
        )


init_db()


def telegram(method, payload=None, files=None, timeout=30):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/{method}"
    response = requests.post(url, json=payload, files=files, timeout=timeout)
    response.raise_for_status()
    return response.json()


def send_message(chat_id, text, keyboard=None, reply_markup=None):
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }
    if reply_markup:
        payload["reply_markup"] = reply_markup
    elif keyboard:
        payload["reply_markup"] = {"inline_keyboard": keyboard}
    telegram("sendMessage", payload)


def main_menu():
    return {
        "keyboard": [
            [{"text": "🧾 رسید جدید"}, {"text": "📊 گزارش ماهانه"}],
            [{"text": "💳 صورت‌حساب کارت"}, {"text": "📈 کنترل کارت"}],
            [{"text": "❓ راهنما"}, {"text": "❌ لغو عملیات"}],
        ],
        "resize_keyboard": True,
        "is_persistent": True,
        "input_field_placeholder": "عکس رسید را ارسال کنید…",
    }


def send_welcome(chat_id):
    send_message(
        chat_id,
        "👋 <b>ایجنت حسابداری Métra Structure</b>\n\n"
        "برای ثبت هزینه، عکس رسید یا فاکتور را بفرست.\n"
        "برای کنترل کارت، PDF صورت‌حساب Mastercard را ارسال کن.\n"
        "همچنین می‌توانی از دکمه‌های پایین صفحه استفاده کنی.",
        reply_markup=main_menu(),
    )


def answer_callback(callback_id):
    try:
        telegram("answerCallbackQuery", {"callback_query_id": callback_id})
    except Exception:
        logger.exception("Unable to answer callback")


def set_session(user_id, step=None, payload=None):
    with db() as connection:
        connection.execute(
            """
            INSERT INTO sessions(user_id, step, payload, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                step=excluded.step,
                payload=excluded.payload,
                updated_at=excluded.updated_at
            """,
            (
                user_id,
                step,
                json.dumps(payload or {}, ensure_ascii=False),
                datetime.utcnow().isoformat(),
            ),
        )


def get_session(user_id):
    with db() as connection:
        row = connection.execute(
            "SELECT step, payload FROM sessions WHERE user_id=?", (user_id,)
        ).fetchone()
    if not row:
        return None, {}
    return row["step"], json.loads(row["payload"])


def is_allowed(user_id):
    return user_id in ALLOWED_USERS


def safe(value):
    return html.escape(str(value or "—"))


def category_keyboard(categories):
    rows = []
    for index, category in enumerate(categories):
        rows.append(
            [{"text": category, "callback_data": f"category:{index}"}]
        )
    rows.append([{"text": "لغو", "callback_data": "cancel"}])
    return rows


def download_receipt(file_id):
    metadata = telegram("getFile", {"file_id": file_id})
    file_path = metadata["result"]["file_path"]
    url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}"
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.content


def download_telegram_file(file_id):
    return download_receipt(file_id)


def _money(text, label, default=0.0):
    match = re.search(
        rf"{label}\s+[-−]?\$?\s*([\d,]+\.\d{{2}})", text, re.IGNORECASE
    )
    return float(match.group(1).replace(",", "")) if match else default


def _parse_date(value):
    cleaned = re.sub(r"\s+", " ", value.strip().replace(",", ""))
    for fmt in ("%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(cleaned.title(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_rbc_statement(pdf_bytes):
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    compact = re.sub(r"[ \t]+", " ", text)

    period = re.search(
        r"STATEMENT FROM\s+([A-Z]{3,9}\s+\d{1,2})\s+TO\s+"
        r"([A-Z]{3,9}\s+\d{1,2},?\s+\d{4})",
        compact,
        re.IGNORECASE,
    )
    if not period:
        raise ValueError("Statement period was not found")
    period_end = _parse_date(period.group(2))
    if not period_end:
        raise ValueError("Statement end date is invalid")
    end_year = int(period_end[:4])
    start_value = f"{period.group(1)} {end_year}"
    period_start = _parse_date(start_value)
    if period_start and period_start > period_end:
        period_start = _parse_date(f"{period.group(1)} {end_year - 1}")

    due = re.search(
        r"Payment due date\s+([A-Z]{3,9}\s+\d{1,2},?\s+\d{4})",
        compact,
        re.IGNORECASE,
    )
    card = re.search(r"\*{2,}\s*(\d{4})\s+-\s+PRIMARY", compact, re.IGNORECASE)

    parsed = {
        "statement_date": period_end,
        "period_start": period_start,
        "period_end": period_end,
        "previous_balance": _money(compact, r"Previous Account Balance"),
        "payments_credits": _money(compact, r"Payments\s*&\s*credits"),
        "purchases_debits": _money(compact, r"Purchases\s*&\s*debits"),
        "cash_advances": _money(compact, r"Cash advances"),
        "interest": _money(compact, r"Interest"),
        "fees": _money(compact, r"Fees"),
        "ending_balance": _money(compact, r"Total Account Balance"),
        "minimum_payment": _money(compact, r"Minimum payment"),
        "payment_due_date": _parse_date(due.group(1)) if due else None,
        "credit_limit": _money(compact, r"Credit limit"),
        "card_last4": card.group(1) if card else None,
    }
    required = ("purchases_debits", "ending_balance", "credit_limit")
    if not all(parsed[name] >= 0 for name in required) or not parsed["credit_limit"]:
        raise ValueError("Statement financial summary is incomplete")
    return parsed


def save_card_statement(user_id, data, digest, source_path):
    with db() as connection:
        connection.execute(
            """
            INSERT INTO card_statements(
                user_id, statement_date, period_start, period_end,
                previous_balance, payments_credits, purchases_debits,
                cash_advances, interest, fees, ending_balance,
                minimum_payment, payment_due_date, credit_limit, card_last4,
                source_hash, source_path, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id, data["statement_date"], data["period_start"],
                data["period_end"], data["previous_balance"],
                data["payments_credits"], data["purchases_debits"],
                data["cash_advances"], data["interest"], data["fees"],
                data["ending_balance"], data["minimum_payment"],
                data["payment_due_date"], data["credit_limit"],
                data["card_last4"], digest, str(source_path),
                datetime.utcnow().isoformat(),
            ),
        )


def card_statement_summary(data):
    utilization = (
        data["ending_balance"] / data["credit_limit"] * 100
        if data["credit_limit"] else 0
    )
    alerts = []
    if data["interest"] > 0:
        alerts.append(f"🔴 بهره ثبت شده: ${data['interest']:,.2f}")
    if data["fees"] > 0:
        alerts.append(f"🔴 کارمزد ثبت شده: ${data['fees']:,.2f}")
    if data["purchases_debits"] > CARD_MONTHLY_BUDGET:
        excess = data["purchases_debits"] - CARD_MONTHLY_BUDGET
        alerts.append(f"🟠 ${excess:,.2f} بالاتر از سقف ماهانه")
    if utilization >= 30:
        alerts.append(f"🟠 استفاده از سقف کارت: {utilization:.1f}٪")
    if not alerts:
        alerts.append("🟢 هشدار مالی جدیدی دیده نشد")
    return (
        f"💳 <b>صورت‌حساب {safe(data['statement_date'])}</b>\n\n"
        f"خریدهای دوره: <b>${data['purchases_debits']:,.2f}</b>\n"
        f"مانده صورت‌حساب: <b>${data['ending_balance']:,.2f}</b>\n"
        f"حداقل پرداخت: ${data['minimum_payment']:,.2f}\n"
        f"سررسید: {safe(data['payment_due_date'])}\n"
        f"بهره: ${data['interest']:,.2f}\n"
        f"کارمزد: ${data['fees']:,.2f}\n"
        f"استفاده از سقف: {utilization:.1f}٪\n\n"
        + "\n".join(alerts)
    )


def card_dashboard(user_id):
    with db() as connection:
        rows = connection.execute(
            """
            SELECT * FROM card_statements
            WHERE user_id=? AND substr(statement_date, 1, 7) >= ?
            ORDER BY statement_date
            """,
            (user_id, CARD_BASELINE_MONTH),
        ).fetchall()
    if not rows:
        return (
            "هنوز صورت‌حسابی برای دوره کنترل جدید ثبت نشده است.\n"
            "PDF صورت‌حساب کارت را ارسال کن."
        )
    spend = sum(row["purchases_debits"] for row in rows)
    interest = sum(row["interest"] for row in rows)
    fees = sum(row["fees"] for row in rows)
    latest = rows[-1]
    utilization = (
        latest["ending_balance"] / latest["credit_limit"] * 100
        if latest["credit_limit"] else 0
    )
    status = "🟢 تحت کنترل"
    if latest["interest"] > 0 or latest["fees"] > 0:
        status = "🔴 نیازمند اقدام"
    elif latest["purchases_debits"] > CARD_MONTHLY_BUDGET or utilization >= 30:
        status = "🟠 نیازمند توجه"
    return (
        f"📈 <b>کنترل کارت شخصی از {safe(CARD_BASELINE_MONTH)}</b>\n\n"
        f"تعداد دوره‌ها: {len(rows)}\n"
        f"مجموع خرید: <b>${spend:,.2f}</b>\n"
        f"میانگین ماهانه: ${spend / len(rows):,.2f}\n"
        f"مجموع بهره: <b>${interest:,.2f}</b>\n"
        f"مجموع کارمزد: ${fees:,.2f}\n"
        f"آخرین مانده: ${latest['ending_balance']:,.2f}\n"
        f"استفاده از سقف: {utilization:.1f}٪\n"
        f"وضعیت: {status}"
    )


def handle_card_document(message):
    user_id = message["from"]["id"]
    chat_id = message["chat"]["id"]
    document = message["document"]
    filename = document.get("file_name", "statement.pdf")
    if document.get("mime_type") != "application/pdf" and not filename.lower().endswith(".pdf"):
        send_message(chat_id, "فقط فایل PDF صورت‌حساب کارت قابل پردازش است.")
        return
    send_message(chat_id, "🔎 صورت‌حساب کارت در حال بررسی است…")
    pdf_bytes = download_telegram_file(document["file_id"])
    digest = hashlib.sha256(pdf_bytes).hexdigest()
    with db() as connection:
        duplicate = connection.execute(
            "SELECT statement_date FROM card_statements WHERE user_id=? AND source_hash=?",
            (user_id, digest),
        ).fetchone()
    if duplicate:
        send_message(chat_id, f"⚠️ صورت‌حساب {safe(duplicate['statement_date'])} قبلاً ثبت شده است.")
        return
    data = parse_rbc_statement(pdf_bytes)
    path = STATEMENT_DIR / f"{user_id}_{data['statement_date']}_{digest[:12]}.pdf"
    path.write_bytes(pdf_bytes)
    save_card_statement(user_id, data, digest, path)
    send_message(chat_id, card_statement_summary(data))


def extract_receipt(image_bytes):
    image_b64 = base64.b64encode(image_bytes).decode("ascii")
    prompt = """
Extract bookkeeping data from this Canadian receipt or invoice.
Return only valid JSON with this exact schema:
{
  "merchant": "string",
  "date": "YYYY-MM-DD",
  "subtotal": 0.0,
  "gst": 0.0,
  "qst": 0.0,
  "total": 0.0,
  "currency": "CAD",
  "description": "short string"
}
Use 0 for taxes that are not shown. Do not estimate a tax that is not visible.
The total must be the amount paid. If the date is unclear, return an empty string.
"""
    response = Anthropic(api_key=ANTHROPIC_API_KEY).messages.create(
        model=os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6"),
        max_tokens=700,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": image_b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )
    raw = "".join(block.text for block in response.content if hasattr(block, "text"))
    raw = raw.replace("```json", "").replace("```", "").strip()
    parsed = json.loads(raw)
    extracted_date = str(parsed.get("date") or "")
    try:
        datetime.strptime(extracted_date, "%Y-%m-%d")
    except ValueError:
        extracted_date = datetime.now().strftime("%Y-%m-%d")
    return {
        "merchant": str(parsed.get("merchant") or "Unknown"),
        "date": extracted_date,
        "subtotal": float(parsed.get("subtotal") or 0),
        "gst": float(parsed.get("gst") or 0),
        "qst": float(parsed.get("qst") or 0),
        "total": float(parsed.get("total") or 0),
        "currency": str(parsed.get("currency") or "CAD").upper(),
        "description": str(parsed.get("description") or ""),
    }


def receipt_summary(data):
    return (
        "🧾 <b>اطلاعات رسید</b>\n\n"
        f"فروشنده: {safe(data['merchant'])}\n"
        f"تاریخ: {safe(data['date'])}\n"
        f"قبل از مالیات: ${data['subtotal']:.2f}\n"
        f"GST: ${data['gst']:.2f}\n"
        f"QST: ${data['qst']:.2f}\n"
        f"<b>جمع: ${data['total']:.2f} {safe(data['currency'])}</b>\n\n"
        "این هزینه مربوط به شرکت است یا شخصی؟"
    )


def save_expense(user_id, payload):
    with db() as connection:
        connection.execute(
            """
            INSERT INTO expenses(
                user_id, merchant, expense_date, subtotal, gst, qst, total,
                currency, description, expense_type, category, project_code,
                receipt_hash, receipt_path, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                payload["merchant"],
                payload["date"],
                payload["subtotal"],
                payload["gst"],
                payload["qst"],
                payload["total"],
                payload["currency"],
                payload["description"],
                payload["expense_type"],
                payload["category"],
                payload.get("project_code"),
                payload["receipt_hash"],
                payload.get("receipt_path"),
                datetime.utcnow().isoformat(),
            ),
        )


def make_excel(user_id, year, month):
    with db() as connection:
        rows = connection.execute(
            """
            SELECT * FROM expenses
            WHERE user_id=? AND substr(expense_date, 1, 7)=?
            ORDER BY expense_date, id
            """,
            (user_id, f"{year:04d}-{month:02d}"),
        ).fetchall()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Expenses"
    headers = [
        "Date",
        "Merchant",
        "Description",
        "Type",
        "Category",
        "Project",
        "Subtotal",
        "GST",
        "QST",
        "Total",
        "Currency",
    ]
    sheet.append(headers)
    navy = "102A43"
    orange = "F5A623"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append(
            [
                row["expense_date"],
                row["merchant"],
                row["description"],
                row["expense_type"],
                row["category"],
                row["project_code"] or "",
                row["subtotal"],
                row["gst"],
                row["qst"],
                row["total"],
                row["currency"],
            ]
        )

    for column in "GHIJ":
        for cell in sheet[column][1:]:
            cell.number_format = '$#,##0.00'
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [12, 24, 32, 12, 28, 16, 14, 12, 12, 14, 10]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Amount"])
    summary.append(["Subtotal", sum(row["subtotal"] for row in rows)])
    summary.append(["GST", sum(row["gst"] for row in rows)])
    summary.append(["QST", sum(row["qst"] for row in rows)])
    summary.append(["Total", sum(row["total"] for row in rows)])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=orange)
    for cell in summary["B"][1:]:
        cell.number_format = '$#,##0.00'
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 16

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, len(rows)


def send_excel(chat_id, user_id, year, month):
    output, count = make_excel(user_id, year, month)
    filename = f"Metra_Bookkeeping_{year:04d}-{month:02d}.xlsx"
    response = requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
        data={
            "chat_id": chat_id,
            "caption": f"گزارش {year:04d}-{month:02d} — {count} سند",
        },
        files={
            "document": (
                filename,
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        timeout=60,
    )
    response.raise_for_status()


def report_keyboard():
    now = datetime.now()
    rows = []
    year, month = now.year, now.month
    for _ in range(12):
        rows.append(
            [
                {
                    "text": f"{year:04d}-{month:02d}",
                    "callback_data": f"report:{year}:{month}",
                }
            ]
        )
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return rows


def handle_photo(message):
    user_id = message["from"]["id"]
    chat_id = message["chat"]["id"]
    send_message(chat_id, "🔎 رسید در حال بررسی است…")
    image_bytes = download_receipt(message["photo"][-1]["file_id"])
    digest = hashlib.sha256(image_bytes).hexdigest()

    with db() as connection:
        duplicate = connection.execute(
            "SELECT id, merchant, total FROM expenses WHERE user_id=? AND receipt_hash=?",
            (user_id, digest),
        ).fetchone()
    if duplicate:
        send_message(
            chat_id,
            f"⚠️ این رسید قبلاً ثبت شده است: {safe(duplicate['merchant'])} — "
            f"${duplicate['total']:.2f}",
        )
        return

    extension = ".jpg"
    receipt_path = RECEIPT_DIR / f"{user_id}_{digest[:20]}{extension}"
    receipt_path.write_bytes(image_bytes)
    data = extract_receipt(image_bytes)
    data["receipt_hash"] = digest
    data["receipt_path"] = str(receipt_path)
    set_session(user_id, "choose_type", data)
    send_message(
        chat_id,
        receipt_summary(data),
        [
            [
                {"text": "🏢 شرکت", "callback_data": "type:company"},
                {"text": "👤 شخصی", "callback_data": "type:personal"},
            ],
            [{"text": "لغو", "callback_data": "cancel"}],
        ],
    )


def handle_callback(callback):
    answer_callback(callback["id"])
    user_id = callback["from"]["id"]
    chat_id = callback["message"]["chat"]["id"]
    if not is_allowed(user_id):
        send_message(chat_id, "⛔️ دسترسی مجاز نیست.")
        return

    action = callback.get("data", "")
    step, payload = get_session(user_id)

    if action == "cancel":
        set_session(user_id)
        send_message(chat_id, "لغو شد.")
        return

    if action.startswith("type:") and step == "choose_type":
        expense_type = action.split(":", 1)[1]
        payload["expense_type"] = expense_type
        categories = (
            COMPANY_CATEGORIES if expense_type == "company" else PERSONAL_CATEGORIES
        )
        payload["categories"] = categories
        set_session(user_id, "choose_category", payload)
        send_message(chat_id, "دسته هزینه را انتخاب کن:", category_keyboard(categories))
        return

    if action.startswith("category:") and step == "choose_category":
        index = int(action.split(":", 1)[1])
        categories = payload.get("categories", [])
        if index >= len(categories):
            raise ValueError("Invalid category")
        payload["category"] = categories[index]
        payload.pop("categories", None)
        if payload.get("expense_type") == "personal":
            payload["project_code"] = None
            save_expense(user_id, payload)
            set_session(user_id)
            send_message(chat_id, "✅ هزینه شخصی با موفقیت ثبت شد.")
            return
        set_session(user_id, "enter_project", payload)
        send_message(
            chat_id,
            "کد پروژه را بنویس؛ مثلاً <b>ODS26-076</b>.\n"
            "اگر هزینه عمومی شرکت است، دکمه زیر را بزن.",
            [[{"text": "هزینه عمومی شرکت", "callback_data": "project:none"}]],
        )
        return

    if action == "project:none" and step == "enter_project":
        payload["project_code"] = None
        save_expense(user_id, payload)
        set_session(user_id)
        send_message(chat_id, "✅ هزینه با موفقیت ثبت شد.")
        return

    if action.startswith("report:"):
        _, year, month = action.split(":")
        send_excel(chat_id, user_id, int(year), int(month))
        return


def handle_message(message):
    user_id = message["from"]["id"]
    chat_id = message["chat"]["id"]
    if not is_allowed(user_id):
        send_message(chat_id, "⛔️ این ربات خصوصی است.")
        return

    text = message.get("text", "").strip()
    step, payload = get_session(user_id)

    if text in {"/start", "/help", "❓ راهنما"}:
        send_welcome(chat_id)
    elif text in {"/cancel", "❌ لغو عملیات"}:
        set_session(user_id)
        send_message(chat_id, "لغو شد.", reply_markup=main_menu())
    elif text in {"/report", "📊 گزارش ماهانه"}:
        send_message(chat_id, "ماه گزارش را انتخاب کن:", report_keyboard())
    elif text in {"/card", "📈 کنترل کارت"}:
        send_message(chat_id, card_dashboard(user_id))
    elif text == "💳 صورت‌حساب کارت":
        set_session(user_id)
        send_message(
            chat_id,
            "📎 فایل PDF صورت‌حساب RBC Mastercard را همین‌جا ارسال کن.",
            reply_markup=main_menu(),
        )
    elif text in {"/new", "🧾 رسید جدید"}:
        set_session(user_id)
        send_message(
            chat_id,
            "📷 عکس واضح رسید یا فاکتور را همین‌جا ارسال کن.",
            reply_markup=main_menu(),
        )
    elif step == "enter_project" and text:
        payload["project_code"] = text.upper()[:50]
        save_expense(user_id, payload)
        set_session(user_id)
        send_message(
            chat_id,
            f"✅ هزینه برای پروژه <b>{safe(payload['project_code'])}</b> ثبت شد.",
        )
    elif message.get("photo"):
        handle_photo(message)
    elif message.get("document"):
        handle_card_document(message)
    else:
        send_message(chat_id, "عکس رسید یا PDF صورت‌حساب کارت را ارسال کن.")


def process_update(update):
    try:
        if "callback_query" in update:
            handle_callback(update["callback_query"])
        elif "message" in update:
            handle_message(update["message"])
    except sqlite3.IntegrityError:
        logger.warning("Duplicate receipt rejected")
        chat_id = (
            update.get("message", {}).get("chat", {}).get("id")
            or update.get("callback_query", {}).get("message", {}).get("chat", {}).get("id")
        )
        if chat_id:
            send_message(chat_id, "⚠️ این رسید قبلاً ثبت شده است.")
    except Exception:
        logger.exception("Update processing failed")
        chat_id = (
            update.get("message", {}).get("chat", {}).get("id")
            or update.get("callback_query", {}).get("message", {}).get("chat", {}).get("id")
        )
        if chat_id:
            send_message(chat_id, "❌ پردازش انجام نشد. دوباره امتحان کن.")


@app.post("/webhook/telegram")
def webhook():
    if WEBHOOK_SECRET:
        supplied = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
        if supplied != WEBHOOK_SECRET:
            return "forbidden", 403
    update = request.get_json(silent=True)
    if update:
        executor.submit(process_update, update)
    return "ok", 200


@app.get("/setup")
def setup():
    if not SETUP_SECRET or request.args.get("key") != SETUP_SECRET:
        return "forbidden", 403
    if not PUBLIC_URL:
        return jsonify({"error": "PUBLIC_URL is required"}), 400
    payload = {
        "url": f"{PUBLIC_URL}/webhook/telegram",
        "allowed_updates": ["message", "callback_query"],
    }
    if WEBHOOK_SECRET:
        payload["secret_token"] = WEBHOOK_SECRET
    webhook_result = telegram("setWebhook", payload)
    commands_result = telegram(
        "setMyCommands",
        {
            "commands": [
                {"command": "start", "description": "شروع و نمایش منوی اصلی"},
                {"command": "new", "description": "ثبت رسید جدید"},
                {"command": "report", "description": "دریافت گزارش ماهانه"},
                {"command": "card", "description": "داشبورد کنترل کارت شخصی"},
                {"command": "help", "description": "راهنمای استفاده"},
                {"command": "cancel", "description": "لغو عملیات جاری"},
            ]
        },
    )
    return jsonify(
        {
            "ok": bool(webhook_result.get("ok") and commands_result.get("ok")),
            "webhook": webhook_result,
            "commands": commands_result,
        }
    )


@app.get("/")
def health():
    return jsonify({"status": "ok", "service": "metra-accounting-bot"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")))
